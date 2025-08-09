import os
from flask import Flask, render_template, request, jsonify, send_file
import io
import csv
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# Load student database
def load_student_database():
    students = {}
    try:
        with open('studentnumbers.txt', 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                parts = line.rsplit('\t', 1) if '\t' in line else line.rsplit(' ', 1)
                if len(parts) == 2:
                    name, number = parts[0].strip(), parts[1].strip()
                    students[name] = number
        return students
    except Exception:
        return {}

student_db = load_student_database()

def create_formatted_excel(data, headers):
    """Create a formatted Excel workbook"""
    wb = Workbook()
    ws = wb.active
    ws.title = "AR Report"
    
    # Add headers
    ws.append(headers)
    
    # Add data rows
    for row in data:
        ws.append([row.get(h, '') for h in headers])
    
    # Style the worksheet
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # Apply styles to headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border
    
    # Apply styles to data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = alignment
            cell.border = thin_border
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    return wb

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process_attendance():
    try:
        # Get uploaded file
        file = request.files.get('attendance_file')
        if not file:
            return jsonify({"error": "No file uploaded"}), 400
        
        # Get selected week
        week = request.form.get('week')
        if not week or not week.isdigit() or int(week) < 1 or int(week) > 16:
            return jsonify({"error": "Please select a valid week (1-16)"}), 400
        
        # Read file based on extension
        filename = file.filename.lower()
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        elif filename.endswith(('.xlsx', '.xls')):
            import pandas as pd
            df = pd.read_excel(file)
            rows = df.to_dict('records')
        else:
            return jsonify({"error": "Unsupported file type. Please upload CSV or Excel file."}), 400
        
        # Process the data
        absent_students = []
        first_module = None
        headers = [
            "Student Number", "Student Name", "Module(s)", "Qualification", 
            "Year", "Week", "Day", "Assessment(s)", 
            "Marks Obtained", "Reason for AR"
        ]
        
        for row in rows:
            if str(row.get('Attendance', '')).lower() == 'absent':
                student_name = str(row.get('Student Name', ''))
                course_code = str(row.get('Course Code', ''))
                section_name = str(row.get('Section Name', ''))
                
                if first_module is None and course_code:
                    first_module = course_code
                
                qualification = section_name.split('(202')[0].strip()
                year_match = re.search(r'^\D*(\d)', course_code)
                year = f"Year{year_match.group(1)}" if year_match else "N/A"
                student_number = student_db.get(student_name, "N/A")
                
                absent_students.append({
                    "Student Number": student_number,
                    "Student Name": student_name,
                    "Module(s)": course_code,
                    "Qualification": qualification,
                    "Year": year,
                    "Week": f"Week{week}",
                    "Day": "N/A",
                    "Assessment(s)": "N/A",
                    "Marks Obtained": "N/A",
                    "Reason for AR": "Class Attendance_007"
                })
        
        # Create formatted Excel file
        wb = create_formatted_excel(absent_students, headers)
        
        # Save to bytes stream
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Generate filename
        current_date = datetime.now().strftime('%Y%m%d')
        filename = f"{first_module}_AR_Report_Week_{week}_{current_date}.xlsx" if first_module else f"AR_Report_Week_{week}_{current_date}.xlsx"
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)
